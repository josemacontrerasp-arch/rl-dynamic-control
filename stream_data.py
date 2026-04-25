"""
Base-case stream conditions for the CO₂-to-methanol flowsheet GNN.

Provides `load_basecase_streams()` — a dict keyed by the 13 edge names in
`flowsheet_graph.STREAMS`, whose values are `StreamState` tuples carrying
(m_dot, T, P, x_CO2, x_H2, x_MeOH, x_H2O, x_inert).  These are used by
`flowsheet_graph.compute_edge_features` to populate the first 8 slots of
the 10-dim edge feature vector (slots 8–9 are structural flags).

Primary source of truth:
    excell/EURECHAEXCELFLOWSHEET.xlsx  — the converged Aspen base case
    (T_rxn=230°C, P_rxn=70 bar, F_H2=2704 kmol/h, P_col=1.56 bar,
     f_purge ≈ 0.005).

Hardcoded fallback values (below) mirror the Excel extract, so this
module works offline and the GNN pipeline never crashes even if the
Excel file is missing or unreadable.  The synthetic-fallback pattern
matches the existing convention in `flowsheet_graph.evaluate_with_surrogates`.

Author: Pepe (Jose Maria Contreras Prada)
Project: EURECHA 2026 Process Design Contest
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Optional


# ─────────────────────────────────────────────────────────────────────
# StreamState — plain record for one edge's stream conditions
# ─────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class StreamState:
    m_dot:   float   # total mass flow (kg/s)
    T:       float   # temperature (°C)
    P:       float   # pressure (bar)
    x_CO2:   float   # mass fraction CO₂
    x_H2:    float   # mass fraction H₂
    x_MeOH:  float   # mass fraction MeOH
    x_H2O:   float   # mass fraction H₂O
    x_inert: float   # mass fraction inerts (CO + N₂ + MEA)

    def as_tuple(self) -> tuple:
        """Return the 8 stream-condition values in edge-feature order."""
        return (
            self.m_dot, self.T, self.P,
            self.x_CO2, self.x_H2, self.x_MeOH, self.x_H2O, self.x_inert,
        )


# ─────────────────────────────────────────────────────────────────────
# Canonical mapping: GNN edge name → Aspen stream name (from Excel)
# ─────────────────────────────────────────────────────────────────────
# Keys MUST match the third element of each tuple in flowsheet_graph.STREAMS.
# Aspen stream names come from `excell/EURECHAEXCELFLOWSHEET.xlsx` row 1.

EDGE_TO_ASPEN: Dict[str, str] = {
    "rich_mea":         "MEAOUT",    # ABSORBER → PUMP (rich MEA from absorber)
    "lean_mea":         "MEALEAN",   # regenerated MEA → absorber (recycle)
    "co2_conc":         "CO2WET",    # STRIPPER → EX5 (concentrated CO₂)
    "co2_comp":         "2",         # COMPCO2 → H2MIXCO2 (compressed CO₂)
    "h2":               "1",         # COMPH2 → H2MIXCO2 (compressed H₂)
    "mixed_feed":       "S6",        # H2MIXCO2 → EX1 (syngas to feed-effluent HX)
    "preheated_feed":   "S8",        # HEAT1 → METREACT (into reactor)
    "reactor_effluent": "S13",       # METREACT → EX1 (reactor outlet)
    "cooled_effluent":  "S23",       # B14 → FLASH1 (into HP flash)
    "fehe_hot":         "S13",       # hot side of feed-effluent HX (energy stream)
    "vapour":           "15",        # FLASH1 → PURGESPT (vapour to splitter)
    "recycle":          "17",        # PURGESPT → H2MIXCO2 (synthesis recycle)
    "crude_meoh":       "S4",        # FLASH2 → COLUMN1 (crude MeOH to distillation)
}


# ─────────────────────────────────────────────────────────────────────
# Hardcoded base case (Aspen converged, extracted from Excel).
# Used as a fallback when the Excel read fails and as the reference
# point against which `compute_edge_features` applies overrides + scaling.
# ─────────────────────────────────────────────────────────────────────

_BASECASE_STREAMS: Dict[str, StreamState] = {
    "rich_mea":         StreamState(163.3901,  50.67,  1.05, 0.0001, 0.0000, 0.0000, 0.9769, 0.0229),
    "lean_mea":         StreamState(157.7283,  40.00,  1.05, 0.0000, 0.0000, 0.0000, 0.7608, 0.2392),
    "co2_conc":         StreamState( 25.2099, 114.85,  1.90, 0.4375, 0.0000, 0.0000, 0.5590, 0.0035),
    "co2_comp":         StreamState( 11.0002,  40.00, 70.00, 0.9992, 0.0000, 0.0000, 0.0008, 0.0000),
    "h2":               StreamState(  1.5141, 138.92, 70.00, 0.0000, 1.0000, 0.0000, 0.0000, 0.0000),
    "mixed_feed":       StreamState( 24.6396,  52.87, 70.00, 0.6962, 0.2744, 0.0123, 0.0019, 0.0153),
    "preheated_feed":   StreamState( 24.6396, 230.00, 70.00, 0.6962, 0.2744, 0.0123, 0.0019, 0.0153),
    "reactor_effluent": StreamState( 24.6400, 230.00, 70.00, 0.2582, 0.2142, 0.3311, 0.1812, 0.0154),
    "cooled_effluent":  StreamState( 24.6400,  40.00, 70.00, 0.2582, 0.2142, 0.3311, 0.1812, 0.0154),
    "fehe_hot":         StreamState( 24.6400, 230.00, 70.00, 0.2582, 0.2142, 0.3311, 0.1812, 0.0154),
    "vapour":           StreamState( 12.1862,  40.00, 70.00, 0.5082, 0.4328, 0.0250, 0.0031, 0.0310),
    "recycle":          StreamState( 12.1253,  40.00, 70.00, 0.5082, 0.4328, 0.0250, 0.0031, 0.0310),
    "crude_meoh":       StreamState( 12.2947,  38.87,  1.56, 0.0034, 0.0000, 0.6368, 0.3598, 0.0000),
}


# Base case scaling anchors — derived from the Aspen base case.
# Used by `flowsheet_graph.compute_edge_features` to scale mass flows.
BASECASE_MEOH_TPH   = 28.25   # t/hr of CH₃OH leaving the reactor at base case
BASECASE_H2_KMOLH   = 2704.0  # kmol/hr of H₂ from PEM at base case (F_H2)
BASECASE_T_RXN_C    = 230.0   # °C
BASECASE_P_RXN_BAR  = 70.0    # bar
BASECASE_P_COL_BAR  = 1.56    # bar
BASECASE_F_PURGE    = 0.005   # purge fraction


# ─────────────────────────────────────────────────────────────────────
# Loader with optional Excel read + graceful fallback
# ─────────────────────────────────────────────────────────────────────

_CACHE: Optional[Dict[str, StreamState]] = None
_CACHE_SOURCE: str = ""  # "excel" or "hardcoded" for diagnostics


def _try_read_excel(path: Path) -> Optional[Dict[str, StreamState]]:
    """Attempt to (re)read the Excel stream table.  Returns None on any
    failure, so callers fall back to the hardcoded values."""
    try:
        import openpyxl  # lazy import — don't pay the cost unless we need it

        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

        def get_row(r: int) -> list:
            return [c.value for c in ws[r]]

        name_row   = get_row(1)
        T_row      = get_row(10)
        P_row      = get_row(11)
        mflow_row  = get_row(42)   # Mass Flows (kg/hr)
        comp_rows = {
            "CO2":      get_row(27),
            "H2":       get_row(28),
            "METHANOL": get_row(29),
            "H2O":      get_row(30),
            "CO":       get_row(31),
            "N2":       get_row(32),
            "MEA":      get_row(34),
        }

        # MW (g/mol) — used to convert kmol/hr → mass fractions
        MW = {"CO2": 44.01, "H2": 2.016, "METHANOL": 32.04,
              "H2O": 18.015, "CO": 28.01, "N2": 28.01, "MEA": 61.08}

        # Build column index → Aspen stream name (col 2+)
        name_to_col: Dict[str, int] = {}
        for i, name in enumerate(name_row):
            if name is None or i < 2:
                continue
            name_to_col[str(name).strip()] = i

        def stream_at(col: int) -> StreamState:
            m_tot_kgh = float(mflow_row[col] or 0)
            m_dot_kgs = m_tot_kgh / 3600.0
            T = float(T_row[col] or 0.0)
            P = float(P_row[col] or 0.0)
            masses = {c: float(rr[col] or 0) * MW[c] for c, rr in comp_rows.items()}
            tot = sum(masses.values())
            if tot <= 0:
                return StreamState(m_dot_kgs, T, P, 0.0, 0.0, 0.0, 0.0, 0.0)
            return StreamState(
                m_dot=m_dot_kgs, T=T, P=P,
                x_CO2 =masses["CO2"] / tot,
                x_H2  =masses["H2"]  / tot,
                x_MeOH=masses["METHANOL"] / tot,
                x_H2O =masses["H2O"] / tot,
                x_inert=(masses["CO"] + masses["N2"] + masses["MEA"]) / tot,
            )

        result: Dict[str, StreamState] = {}
        for edge, aspen in EDGE_TO_ASPEN.items():
            if aspen not in name_to_col:
                warnings.warn(f"[stream_data] Aspen stream {aspen!r} not found in Excel; "
                              f"using hardcoded base case for edge {edge!r}.")
                result[edge] = _BASECASE_STREAMS[edge]
            else:
                result[edge] = stream_at(name_to_col[aspen])
        return result
    except Exception as e:
        warnings.warn(f"[stream_data] Excel read failed ({type(e).__name__}: {e}); "
                      f"falling back to hardcoded base case.")
        return None


def load_basecase_streams(
    excel_path: Optional[Path] = None,
    force_reload: bool = False,
) -> Dict[str, StreamState]:
    """Return the dict of base-case stream states, keyed by edge name.

    On first call, tries to read `excell/EURECHAEXCELFLOWSHEET.xlsx`; if
    that fails (file missing, openpyxl not installed, etc.) falls back
    to the hardcoded constants in this module.  Subsequent calls return
    the cached result unless `force_reload=True`.
    """
    global _CACHE, _CACHE_SOURCE
    if _CACHE is not None and not force_reload:
        return _CACHE

    if excel_path is None:
        # Default: project-relative .../EURECHA/excell/EURECHAEXCELFLOWSHEET.xlsx
        here = Path(__file__).resolve().parent
        excel_path = here.parent / "excell" / "EURECHAEXCELFLOWSHEET.xlsx"

    if excel_path.exists():
        read = _try_read_excel(excel_path)
        if read is not None:
            _CACHE = read
            _CACHE_SOURCE = "excel"
            return _CACHE

    _CACHE = dict(_BASECASE_STREAMS)
    _CACHE_SOURCE = "hardcoded"
    return _CACHE


def basecase_source() -> str:
    """Return 'excel' or 'hardcoded' depending on where the cached base
    case came from.  Returns '' if `load_basecase_streams` hasn't been
    called yet.  Useful for test assertions and diagnostics."""
    return _CACHE_SOURCE


def reset_cache() -> None:
    """Clear the module-level cache.  Tests use this to force reload."""
    global _CACHE, _CACHE_SOURCE
    _CACHE = None
    _CACHE_SOURCE = ""


# ─────────────────────────────────────────────────────────────────────
# Stream-condition array builder (plain numpy — no torch dependency)
# ─────────────────────────────────────────────────────────────────────
#
# The actual edge-feature function in flowsheet_graph.py is a thin
# wrapper around this.  Keeping the physics-heavy logic here lets us
# unit-test it without importing torch.

# Edge order — MUST match `flowsheet_graph.STREAMS` element-for-element.
EDGE_NAMES_ORDERED = [
    "rich_mea",
    "lean_mea",
    "co2_conc",
    "co2_comp",
    "h2",
    "mixed_feed",
    "preheated_feed",
    "reactor_effluent",
    "cooled_effluent",
    "fehe_hot",
    "vapour",
    "recycle",
    "crude_meoh",
]

# Edge-index sets used for structural flags + overrides.
RECYCLE_EDGE_IDX = {1, 11}     # lean_mea, recycle
ENERGY_EDGE_IDX  = {9}         # fehe_hot
_T_RXN_EDGES = {"preheated_feed", "reactor_effluent", "fehe_hot"}
_P_RXN_EDGES = {
    "h2", "co2_comp", "mixed_feed", "preheated_feed",
    "reactor_effluent", "fehe_hot", "cooled_effluent",
    "vapour", "recycle",
}
_P_COL_EDGES = {"crude_meoh"}

EDGE_FEATURE_DIM = 10          # mirrors flowsheet_graph.EDGE_FEATURE_DIM


def build_edge_feature_array(
    node_params: dict,
    surrogate_out: Optional[dict] = None,
) -> "np.ndarray":
    """Return a (13, 10) stream-condition + flag array for one sample.

    See ``flowsheet_graph.compute_edge_features`` for the semantics —
    this function contains the actual logic; the other is a thin
    wrapper used by the GNN pipeline.

    Errors are *not* swallowed here: the wrapper catches them and
    returns a structural-flag-only fallback.  Tests that exercise the
    fallback path monkeypatch this module's ``load_basecase_streams``
    to raise.
    """
    import numpy as np

    base = load_basecase_streams()

    # Node-param overrides — default to Aspen base case if keys missing.
    T_rxn = float(node_params.get(6, [BASECASE_T_RXN_C, BASECASE_P_RXN_BAR])[0]) \
        if 6 in node_params else BASECASE_T_RXN_C
    P_rxn = float(node_params.get(6, [BASECASE_T_RXN_C, BASECASE_P_RXN_BAR])[1]) \
        if 6 in node_params else BASECASE_P_RXN_BAR
    P_col = float(node_params.get(10, [0, 0, BASECASE_P_COL_BAR])[2]) \
        if 10 in node_params else BASECASE_P_COL_BAR

    # Mass-flow scaling from the MeOH surrogate (when available).
    meoh_tph = None
    if surrogate_out is not None:
        meoh_tph = surrogate_out.get("_meoh_tph")
    m_scale = float(meoh_tph) / BASECASE_MEOH_TPH \
        if (meoh_tph is not None and meoh_tph > 0) else 1.0

    arr = np.zeros((len(EDGE_NAMES_ORDERED), EDGE_FEATURE_DIM), dtype=float)
    for i, name in enumerate(EDGE_NAMES_ORDERED):
        s = base[name]
        T = T_rxn if name in _T_RXN_EDGES else s.T
        if name in _P_COL_EDGES:
            P = P_col
        elif name in _P_RXN_EDGES:
            P = P_rxn
        else:
            P = s.P
        arr[i, 0] = s.m_dot * m_scale
        arr[i, 1] = T
        arr[i, 2] = P
        arr[i, 3] = s.x_CO2
        arr[i, 4] = s.x_H2
        arr[i, 5] = s.x_MeOH
        arr[i, 6] = s.x_H2O
        arr[i, 7] = s.x_inert

    # Structural flags (indices 8, 9).
    for i in RECYCLE_EDGE_IDX:
        arr[i, 8] = 1.0
    for i in ENERGY_EDGE_IDX:
        arr[i, 9] = 1.0
    return arr
